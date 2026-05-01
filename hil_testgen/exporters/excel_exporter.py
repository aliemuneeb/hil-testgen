"""
Excel exporter — saves generated test cases to .xlsx
Three tabs: Generated, Needs Review, Summary
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)


class ExcelExporter:
    """
    Exports test cases to a professional .xlsx file.

    Tab 1: Generated Test Cases  ← ready to use
    Tab 2: Needs Review          ← skipped with reasons
    Tab 3: Summary               ← stats at a glance
    """

    # Colors
    COLOR_HIGH    = "E8F5E9"  # light green
    COLOR_REVIEW  = "FFF8E1"  # light amber
    COLOR_LOW     = "FFEBEE"  # light red
    COLOR_HEADER  = "1A237E"  # dark blue
    COLOR_SKIP_H  = "B71C1C"  # dark red
    COLOR_SUMMARY = "E3F2FD"  # light blue
    COLOR_WHITE   = "FFFFFF"

    # Column definitions for generated test cases
    TC_COLUMNS = [
        ("TC ID",             "tc_id",             20),
        ("Requirement ID",    "req_id",             18),
        ("Title",             "title",              30),
        ("Objective",         "objective",          40),
        ("Preconditions",     "preconditions",      35),
        ("Input Signals",     "input_signals",      35),
        ("Test Steps",        "test_steps",         45),
        ("Expected Output",   "expected_output",    35),
        ("Pass Criteria",     "pass_criteria",      35),
        ("Fail Criteria",     "fail_criteria",      35),
        ("Priority",          "priority",           12),
        ("Test Type",         "test_type",          15),
        ("Confidence",        "confidence",         13),
        ("Confidence Score",  "confidence_score",   18),
        ("Notes",             "notes",              35),
        ("Confidence Notes",  "confidence_notes",   40),
    ]

    # Column definitions for skipped requirements
    SKIP_COLUMNS = [
        ("Requirement ID",  "req_id",   18),
        ("Reason",          "reason",   40),
        ("Missing",         "missing",  50),
    ]

    def __init__(
        self,
        test_cases : list,
        skipped    : list,
        output_path: Path
    ):
        self.test_cases  = test_cases
        self.skipped     = skipped
        self.output_path = output_path

    def export(self):
        """Build and save the Excel workbook."""
        wb = Workbook()

        # ── Tab 1: Generated test cases ──────────────────────────────────
        ws_tc = wb.active
        ws_tc.title = "Generated Test Cases"
        self._write_tc_sheet(ws_tc)

        # ── Tab 2: Needs review ──────────────────────────────────────────
        ws_skip = wb.create_sheet("Needs Review")
        self._write_skip_sheet(ws_skip)

        # ── Tab 3: Summary ───────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_sum)

        wb.save(self.output_path)

    # ── Sheet writers ────────────────────────────────────────────────────

    def _write_tc_sheet(self, ws):
        """Write generated test cases sheet."""

        # Header row
        headers = [col[0] for col in self.TC_COLUMNS]
        self._write_header_row(ws, headers, self.COLOR_HEADER)

        # Data rows
        for row_idx, tc in enumerate(self.test_cases, start=2):
            confidence = tc.get("confidence", "REVIEW")
            row_color  = self._confidence_color(confidence)

            for col_idx, (_, field, _) in enumerate(self.TC_COLUMNS, start=1):
                value = tc.get(field, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=str(value))
                self._style_data_cell(cell, row_color)

                # Special formatting for confidence and priority columns
                if field == "confidence":
                    cell.font = Font(
                        bold=True,
                        color=self._confidence_font_color(confidence)
                    )
                if field == "priority":
                    cell.font = Font(
                        bold=True,
                        color=self._priority_font_color(str(value))
                    )

        # Column widths
        for col_idx, (_, _, width) in enumerate(self.TC_COLUMNS, start=1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

    def _write_skip_sheet(self, ws):
        """Write skipped requirements sheet."""

        if not self.skipped:
            ws.cell(row=1, column=1, value="No requirements were skipped.")
            return

        headers = [col[0] for col in self.SKIP_COLUMNS]
        self._write_header_row(ws, headers, self.COLOR_SKIP_H)

        for row_idx, skip in enumerate(self.skipped, start=2):
            for col_idx, (_, field, _) in enumerate(self.SKIP_COLUMNS, start=1):
                value = skip.get(field, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=str(value))
                self._style_data_cell(cell, self.COLOR_REVIEW)

        for col_idx, (_, _, width) in enumerate(self.SKIP_COLUMNS, start=1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, ws):
        """Write summary statistics sheet."""

        total     = len(self.test_cases) + len(self.skipped)
        generated = len(self.test_cases)
        skipped   = len(self.skipped)

        high_conf   = sum(
            1 for tc in self.test_cases
            if tc.get("confidence") == "HIGH"
        )
        review_conf = sum(
            1 for tc in self.test_cases
            if tc.get("confidence") == "REVIEW"
        )
        low_conf    = sum(
            1 for tc in self.test_cases
            if tc.get("confidence") == "LOW"
        )

        rows = [
            ("hil-testgen Report", ""),
            ("", ""),
            ("OVERVIEW", ""),
            ("Total Requirements Found",  total),
            ("Test Cases Generated",      generated),
            ("Requirements Skipped",      skipped),
            ("", ""),
            ("CONFIDENCE BREAKDOWN", ""),
            ("High Confidence",    high_conf),
            ("Needs Review",       review_conf),
            ("Low Confidence",     low_conf),
            ("", ""),
            ("NOTES", ""),
            (
                "High Confidence",
                "Requirement was complete — test case ready to use"
            ),
            (
                "Needs Review",
                "Some fields were missing — verify before use"
            ),
            (
                "Low Confidence",
                "Significant info missing — AI inferred values"
            ),
            ("", ""),
            (
                "Disclaimer",
                "All generated test cases must be reviewed by a "
                "qualified engineer before use in validation."
            ),
        ]

        for row_idx, (label, value) in enumerate(rows, start=1):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            value_cell = ws.cell(row=row_idx, column=2, value=value)

            # Title row
            if label == "hil-testgen Report":
                label_cell.font = Font(
                    bold=True, size=14, color=self.COLOR_HEADER
                )

            # Section headers
            elif label in ["OVERVIEW", "CONFIDENCE BREAKDOWN", "NOTES"]:
                label_cell.font = Font(bold=True, color=self.COLOR_HEADER)
                label_cell.fill = PatternFill(
                    "solid", fgColor=self.COLOR_SUMMARY
                )
                value_cell.fill = PatternFill(
                    "solid", fgColor=self.COLOR_SUMMARY
                )

            # Data rows
            elif label and value != "":
                label_cell.font = Font(bold=False)
                value_cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60

    # ── Styling helpers ──────────────────────────────────────────────────

    def _write_header_row(self, ws, headers: list, color: str):
        """Write a styled header row."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = Font(bold=True, color=self.COLOR_WHITE, size=11)
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border    = self._thin_border()
        ws.row_dimensions[1].height = 30

    def _style_data_cell(self, cell, row_color: str):
        """Apply standard styling to a data cell."""
        cell.fill      = PatternFill("solid", fgColor=row_color)
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )
        cell.border    = self._thin_border()
        ws = cell.parent
        ws.row_dimensions[cell.row].height = 60

    def _thin_border(self) -> Border:
        """Return a thin border for cells."""
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _confidence_color(self, confidence: str) -> str:
        """Return row background color based on confidence level."""
        return {
            "HIGH"  : self.COLOR_HIGH,
            "REVIEW": self.COLOR_REVIEW,
            "LOW"   : self.COLOR_LOW,
        }.get(confidence, self.COLOR_WHITE)

    def _confidence_font_color(self, confidence: str) -> str:
        """Return font color based on confidence level."""
        return {
            "HIGH"  : "1B5E20",  # dark green
            "REVIEW": "E65100",  # dark amber
            "LOW"   : "B71C1C",  # dark red
        }.get(confidence, "000000")

    def _priority_font_color(self, priority: str) -> str:
        """Return font color based on priority."""
        return {
            "HIGH"  : "B71C1C",  # dark red
            "MEDIUM": "E65100",  # dark amber
            "LOW"   : "1B5E20",  # dark green
        }.get(priority, "000000")
